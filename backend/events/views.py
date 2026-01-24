from django.utils import timezone
from datetime import timedelta, datetime
from django.http import JsonResponse, HttpRequest
import json, re
import time
from events.communication import send_emails
from .models import (
    Event,
    Participant,
    ParticipantLog,
    Website,
    BlockedHost,
    ParticipantEvent,
    EventConsent,
)
from .s3_service import s3_service

from django.views.decorators.csrf import csrf_exempt
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Q
from django.db.models import Max
from zoneinfo import ZoneInfo
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q
from django.db.models.deletion import RestrictedError
from django.core.cache import cache
from authentication.models import CustomUser, UserRole
from authentication.views import verify_token, get_user_data
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from io import BytesIO
from django.http import HttpResponse
from authentication.utils import jwt_required
from django.views.decorators.http import require_POST, require_GET
import logging
from django.utils import timezone
from django.db import transaction
logger = logging.getLogger(__name__)
from behavior_analysis.models import AnalisisComportamiento
from events.tasks import delete_event_media_from_s3

EVENT_EXPIRATION_DAYS = 182
MONITORING_UPLOAD_GRACE_SECONDS = 300

# Funciones


def check_event_time(event):
    now = timezone.now()
    if event.start_date and event.end_date:
        return event.start_date <= now <= event.end_date
    return False


def validate_event_access(participant_event, now):
    """
    Valida si un participante puede acceder al evento según las reglas de negocio
    """
    event = participant_event.event

    # Verificar si el participante está bloqueado por el administrador
    if participant_event.is_blocked:
        return {
            "allowed": False,
            "reason": "Tu acceso ha sido bloqueado por el administrador.",
            "monitoring_allowed": False,
            "is_first_connection": False,
        }

    # Determinar si es primera conexión basado en si ha monitoreado antes
    is_first_connection = participant_event.monitoring_sessions_count == 0
    has_connected_before = not is_first_connection

    # Regla 0: Verificar si ya consumió todo su tiempo
    total_time_seconds = participant_event.get_total_monitoring_time()
    event_duration_seconds = event.duration * 60

    # Dar un margen de tolerancia de 30 segundos
    if total_time_seconds >= (event_duration_seconds + 30):
        return {
            "allowed": False,
            "reason": "Has consumido todo el tiempo disponible para este evento.",
            "monitoring_allowed": False,
            "is_first_connection": is_first_connection,
        }

    # Regla 1: Nadie puede entrar después de end_date (fecha final absoluta)
    if now > event.end_date:
        return {
            "allowed": False,
            "reason": "El evento ha finalizado.",
            "monitoring_allowed": False,
            "is_first_connection": is_first_connection,
        }

    # Regla 2: Primera conexión no permitida después de close_date
    if is_first_connection and now > event.close_date:
        return {
            "allowed": False,
            "reason": "El período de ingreso al evento ha terminado.",
            "monitoring_allowed": False,
            "is_first_connection": is_first_connection,
        }

    # Regla 3: Conexiones previas permitidas hasta end_date
    if has_connected_before and now <= event.end_date:
        return {
            "allowed": True,
            "reason": "Acceso permitido - participante previamente conectado",
            "monitoring_allowed": True,  # Permitir monitoreo hasta end_date si ya se conectó antes
            "is_first_connection": is_first_connection,
        }

    # Regla 4: Primera conexión dentro del período normal
    if is_first_connection and now <= event.close_date:
        return {
            "allowed": True,
            "reason": "Acceso permitido - dentro del período de conexión",
            "monitoring_allowed": True,
            "is_first_connection": is_first_connection,
        }

    # Caso por defecto (no debería llegar aquí)
    return {
        "allowed": False,
        "reason": "Acceso no permitido",
        "monitoring_allowed": False,
        "is_first_connection": is_first_connection,
    }


def is_valid_domain(domain):
    # Expresión regular básica para dominios (no URLs completas)
    pattern = r"^(?!\-)([A-Za-z0-9\-]{1,63}(?<!\-)\.)+[A-Za-z]{2,}$"
    return re.match(pattern, domain) is not None


def _extract_s3_key(value):
    if not value:
        return None
    if "amazonaws.com" in value:
        return value.split(".amazonaws.com/")[-1].split("?")[0]
    return value


def _collect_event_media_keys(event_id):
    log_keys = (
        ParticipantLog.objects.filter(participant_event__event_id=event_id)
        .exclude(url__isnull=True)
        .exclude(url__exact="")
        .values_list("url", flat=True)
    )
    analysis_keys = (
        AnalisisComportamiento.objects.filter(participant_event__event_id=event_id)
        .exclude(video_link__isnull=True)
        .exclude(video_link__exact="")
        .values_list("video_link", flat=True)
    )
    keys = set()
    for key in log_keys:
        normalized = _extract_s3_key(key)
        if normalized:
            keys.add(normalized)
    for key in analysis_keys:
        normalized = _extract_s3_key(key)
        if normalized:
            keys.add(normalized)
    return keys


def verify_event_key(request):
    authorization = request.headers.get("Authorization", "")

    if not authorization.startswith("Bearer "):
        return JsonResponse({"error": "Invalid format authorization"}, status=401)
    parts = authorization.split(" ", 1)
    if len(parts) != 2 or not parts[1]:
        return JsonResponse({"error": "Invalid format authorization"}, status=401)

    event_key = parts[1]
    try:
        participant_event = ParticipantEvent.objects.select_related(
            "participant", "event"
        ).get(event_key=event_key)
        participant = participant_event.participant
        event = participant_event.event

        from django.utils import timezone

        now = timezone.now()

        # Validación básica de tiempo del evento (start_date - 1min hasta end_date)
        dateIsValid = check_event_time(event)

        # Validaciones avanzadas de acceso
        access_validation = validate_event_access(participant_event, now)

        if not access_validation["allowed"]:
            return JsonResponse(
                {
                    "isValid": False,
                    "error": access_validation["reason"],
                    "dateIsValid": False,
                    "specificError": True,  # Indica que es un error específico, no genérico
                },
                status=403,
            )

        # Verificar si existe consentimiento informado para este participante y evento
        consent_exists = EventConsent.objects.filter(
            participant=participant, event=event
        ).exists()

        # Si no existe consentimiento, requerir aceptación antes de continuar
        if not consent_exists:
            return JsonResponse(
                {
                    "isValid": True,
                    "dateIsValid": dateIsValid,
                    "consentRequired": True,
                    "participant": {
                        "id": participant.id,
                        "name": participant.name,
                        "email": participant.email,
                    },
                    "event": {
                        "id": event.id,
                        "name": event.name,
                        "description": event.description,
                        "duration": event.duration,
                    },
                    "message": "Debe aceptar el consentimiento informado antes de continuar",
                },
                status=200,
            )

        # Obtener información del tiempo de monitoreo desde ParticipantEvent
        connection_info = {
            "totalTime": participant_event.get_total_monitoring_time(),
            "totalTimeSeconds": participant_event.get_total_monitoring_seconds(),
            "isActive": participant_event.is_monitoring,
            "eventDuration": event.duration,
            "monitoringAllowed": access_validation["monitoring_allowed"],
            "sessionCount": participant_event.monitoring_sessions_count,
            "isFirstConnection": access_validation["is_first_connection"],
        }

        return JsonResponse(
            {
                "isValid": True,
                "dateIsValid": dateIsValid,
                "consentRequired": False,
                "participant": {
                    "name": participant.name,
                    "email": participant.email,
                    "monitoring_total_duration": participant_event.monitoring_total_duration,
                },
                "event": {
                    "name": event.name,
                    "id": event.id,
                    "duration": event.duration,
                },
                "connectionInfo": connection_info,
            }
        )
    except ParticipantEvent.DoesNotExist:
        return JsonResponse({"isValid": False}, status=404)


def get_participant(event_key):
    try:
        participant_event = ParticipantEvent.objects.select_related("participant").get(
            event_key=event_key
        )
        return participant_event.participant
    except ParticipantEvent.DoesNotExist:
        return None


def get_participant_event(event_key):
    """Devuelve el ParticipantEvent asociado al event_key o None si no existe.

    Mantener esta función separada de `get_participant` para no romper código
    que pueda depender todavía en el objeto Participant.
    """
    try:
        participant_event = ParticipantEvent.objects.select_related(
            "participant", "event"
        ).get(event_key=event_key)
        return participant_event
    except ParticipantEvent.DoesNotExist:
        return None


def _allow_upload_after_block(participant_event, now=None):
    if now is None:
        now = timezone.now()

    if getattr(participant_event, "is_monitoring", False):
        return True

    if getattr(participant_event, "monitoring_sessions_count", 0) <= 0:
        return False

    event = getattr(participant_event, "event", None)
    if event and event.end_date:
        if now <= event.end_date + timedelta(seconds=MONITORING_UPLOAD_GRACE_SECONDS):
            return True

    if not getattr(participant_event, "is_blocked", False):
        return False

    last_change = getattr(participant_event, "monitoring_last_change", None)
    if not last_change:
        return False

    delta_seconds = (now - last_change).total_seconds()
    return 0 <= delta_seconds <= MONITORING_UPLOAD_GRACE_SECONDS


def _validate_participant_fields(
    first_name: str,
    last_name: str,
    email: str,
    seen_emails: set,
    participant_id: int = None,
    ids_in_excel: set = None,
):
    """Valida campos de participante.

    Args:
        first_name: Nombre del participante
        last_name: Apellidos del participante
        email: Email del participante
        seen_emails: Set de emails ya vistos en el lote actual
        participant_id: ID del participante (para actualizaciones, excluye self de validación de duplicados)
        ids_in_excel: Set de IDs que están siendo actualizados en este lote

    Returns:
        Tupla (first_name, last_name, email, errors)
    """
    errors = []
    fn = (first_name or "").strip()
    ln = (last_name or "").strip()
    em = (email or "").strip().lower()

    if not fn:
        errors.append("Nombre requerido")
    if not ln:
        errors.append("Apellidos requeridos")
    if not em:
        errors.append("Email requerido")
    else:
        try:
            validate_email(em)
        except ValidationError:
            errors.append("Email inválido")

    # Duplicado dentro del archivo (o del payload)
    if em:
        if em in seen_emails:
            errors.append("Email duplicado en el archivo")
        else:
            seen_emails.add(em)

    # Duplicado en BD (excluir el propio participante si es actualización)
    if em:
        qs = Participant.objects.filter(email__iexact=em)
        if participant_id:
            qs = qs.exclude(id=participant_id)

        # Obtener conflictos potenciales
        conflicts = list(qs.values_list("id", flat=True))

        if conflicts:
            # Si hay conflictos, verificar si todos los conflictos están siendo actualizados en este Excel
            # Si el ID conflictivo está en el Excel, asumimos que su correo será sobrescrito (o es el mismo, validado por seen_emails)

            # Si ids_in_excel es None, comportamiento antiguo (error directo)
            if ids_in_excel is None:
                errors.append("Email ya existe en el sistema")
            else:
                # Verificar si hay algún conflicto que NO esté en el Excel
                # Si hay un conflicto con un ID que NO está en el Excel, es un error real.
                # NOTA: Si el ID no está en el Excel, será eliminado, por lo que el correo quedará libre.
                # Por lo tanto, si estamos en modo "full sync" (ids_in_excel != None),
                # NO deberíamos marcar error si el conflicto es con alguien que no está en el Excel.

                # Conflictos reales: IDs que existen en BD, tienen ese email, Y TAMBIÉN están en el Excel (pero con otro ID, obvio)
                # Si el conflicto está en el Excel, significa que ese usuario está siendo actualizado.
                # Si ese usuario actualizado MANTIENE su email, entonces seen_emails lo habría detectado (duplicado en Excel).
                # Si ese usuario actualizado CAMBIA su email, entonces libera este email.

                # Por lo tanto, la única validación necesaria es seen_emails (duplicados dentro del Excel).
                # Los conflictos con la BD se resuelven solos:
                # 1. Si el conflicto NO está en Excel -> Se elimina -> Email libre.
                # 2. Si el conflicto SÍ está en Excel -> Se actualiza.
                #    a. Si mantiene email -> Duplicado en Excel -> Error en seen_emails.
                #    b. Si cambia email -> Email libre.

                # Conclusión: En modo full sync, NO necesitamos validar contra BD, solo contra seen_emails.
                pass

    return fn, ln, em, errors


def handle_event_participants(event, participants_payload):
    """
    Sin crear ni eliminar objetos Participant.
    Solo sincroniza la relación M2M Event <-> Participant
    en base a los items del payload con selected=True.

    participants_payload: lista de dicts del tipo:
        { "email": "a@b.com", "name": "...", "selected": true/false }
    """

    # Normalizar correos del payload
    def _norm(email):
        return (email or "").strip().lower()

    selected_emails = {
        _norm(item.get("email"))
        for item in (participants_payload or [])
        if item.get("selected", False) and item.get("email")
    }

    # Estado actual en BD
    current_participants = list(event.participants.all())
    current_emails = {p.email.lower() for p in current_participants}

    # A quién agregar / quitar (solo por email existente en BD)
    emails_to_add = selected_emails - current_emails
    emails_to_remove = current_emails - selected_emails

    # Traer solo los Participant existentes que corresponden a emails_to_add
    participants_to_add = list(
        Participant.objects.filter(email__in=list(emails_to_add))
    )

    # Map rápido por email
    current_by_email = {p.email.lower(): p for p in current_participants}

    with transaction.atomic():
        # Agregar asociaciones nuevas
        for participant in participants_to_add:
            ParticipantEvent.objects.get_or_create(event=event, participant=participant)

        # Quitar asociaciones eliminando el registro intermedio
        for email in emails_to_remove:
            participant = current_by_email.get(email)
            if participant:
                ParticipantEvent.objects.filter(
                    event=event, participant=participant
                ).delete()

    # Métricas útiles para logs/debug
    added_count = len(participants_to_add)
    removed_count = len(emails_to_remove)
    skipped_emails = sorted(
        selected_emails
        - {p.email.lower() for p in participants_to_add}
        - current_emails
    )
    # `skipped_emails` = correos marcados como selected que NO existen en Participant,
    # y por política no se crean aquí.

    print(
        f"[handle_event_participants] Added: {added_count}, Removed: {removed_count}, "
        f"Skipped (not found in Participant): {len(skipped_emails)} -> {skipped_emails}"
    )

    return {
        "added": added_count,
        "removed": removed_count,
        "skipped_not_found": skipped_emails,
    }


# Endpoints app escritorio


@require_POST
def log_participant_http_event(request: HttpRequest):

    try:
        data = json.loads(request.body)
        auth = request.headers.get("Authorization", "")
        parts = auth.split(" ", 1)
        if len(parts) != 2 or not parts[1]:
            return JsonResponse({"error": "Invalid format authorization"}, status=401)
        event_key = parts[1]
        participant_event = get_participant_event(event_key)
        if not participant_event:
            return JsonResponse({"error": "ParticipantEvent not found"}, status=404)

        now = timezone.now()
        if not _allow_upload_after_block(participant_event, now):
            return JsonResponse({"error": "Monitoring not started"}, status=403)

        log_type = data.get("type", "http")
        ParticipantLog.objects.create(
            name=log_type, message=data["uri"], participant_event=participant_event
        )
        return JsonResponse({"status": "success"})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_POST
def presign_participant_screen_upload(request: HttpRequest):
    try:
        auth = request.headers.get("Authorization", "")
        parts = auth.split(" ", 1)
        if len(parts) != 2 or not parts[1]:
            return JsonResponse({"error": "Invalid format authorization"}, status=401)
        event_key = parts[1]
        participant_event = get_participant_event(event_key)
        if not participant_event:
            return JsonResponse({"error": "ParticipantEvent not found"}, status=404)

        if not _allow_upload_after_block(participant_event, timezone.now()):
            return JsonResponse({"error": "Monitoring not started"}, status=403)

        upload_result = s3_service.generate_presigned_upload(
            participant_event.id, media_type="screen", timestamp=timezone.now()
        )

        if upload_result["success"]:
            return JsonResponse(
                {
                    "status": "success",
                    "s3_key": upload_result["key"],
                    "upload_url": upload_result["upload_url"],
                    "headers": upload_result["headers"],
                }
            )
        return JsonResponse(
            {"error": f"Failed to presign upload: {upload_result['error']}"},
            status=500,
        )

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_POST
def presign_participant_media_upload(request: HttpRequest):
    try:
        auth = request.headers.get("Authorization", "")
        parts = auth.split(" ", 1)
        if len(parts) != 2 or not parts[1]:
            return JsonResponse({"error": "Invalid format authorization"}, status=401)
        event_key = parts[1]
        participant_event = get_participant_event(event_key)
        if not participant_event:
            return JsonResponse({"error": "ParticipantEvent not found"}, status=404)

        if not _allow_upload_after_block(participant_event, timezone.now()):
            return JsonResponse({"error": "Monitoring not started"}, status=403)

        media_type = "video"
        try:
            data = json.loads(request.body or "{}")
            if isinstance(data, dict):
                media_type = data.get("media_type", media_type)
        except json.JSONDecodeError:
            pass

        upload_result = s3_service.generate_presigned_upload(
            participant_event.id, media_type=media_type, timestamp=timezone.now()
        )

        if upload_result["success"]:
            return JsonResponse(
                {
                    "status": "success",
                    "s3_key": upload_result["key"],
                    "upload_url": upload_result["upload_url"],
                    "headers": upload_result["headers"],
                }
            )
        return JsonResponse(
            {"error": f"Failed to presign upload: {upload_result['error']}"},
            status=500,
        )

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_POST
def log_participant_screen_event(request: HttpRequest):

    try:
        auth = request.headers.get("Authorization", "")
        parts = auth.split(" ", 1)
        if len(parts) != 2 or not parts[1]:
            return JsonResponse({"error": "Invalid format authorization"}, status=401)
        event_key = parts[1]
        participant_event = get_participant_event(event_key)
        if not participant_event:
            return JsonResponse({"error": "ParticipantEvent not found"}, status=404)

        if not getattr(participant_event, "is_monitoring", False):
            return JsonResponse({"error": "Monitoring not started"}, status=403)

        now = timezone.now()
        s3_key = None
        monitor_name = "Unknown Monitor"

        if "screenshot" in request.FILES:
            file = request.FILES["screenshot"]
            monitor_name = request.POST.get("monitor_name", "Unknown Monitor")

            # Subir archivo a S3
            upload_result = s3_service.upload_media_fragment(
                file,
                participant_event.id,
                media_type="screen",
                timestamp=now,
            )

            if upload_result["success"]:
                s3_key = upload_result["key"]
                presigned_url = upload_result.get("presigned_url")
            else:
                return JsonResponse(
                    {"error": f"Failed to upload screenshot: {upload_result['error']}"},
                    status=500,
                )
        else:
            try:
                data = json.loads(request.body or "{}")
            except json.JSONDecodeError:
                data = {}

            s3_key = data.get("s3_key") or request.POST.get("s3_key")
            monitor_name = data.get("monitor_name") or request.POST.get(
                "monitor_name", "Unknown Monitor"
            )
            presigned_url = (
                s3_service.generate_presigned_url(s3_key) if s3_key else None
            )

        if not s3_key:
            return JsonResponse(
                {"error": "Missing screenshot file or s3_key"},
                status=400,
            )

        # Guardar el log con la URL de S3
        ParticipantLog.objects.create(
            name="screen",
            url=s3_key,  # guardamos la key en el campo url
            message=f"{monitor_name}",
            participant_event=participant_event,
        )
        return JsonResponse(
            {
                "status": "success",
                "s3_key": s3_key,
                "url": presigned_url,
            }
        )

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_POST
def log_participant_audio_video_event(request: HttpRequest):

    try:
        auth = request.headers.get("Authorization", "")
        parts = auth.split(" ", 1)
        if len(parts) != 2 or not parts[1]:
            return JsonResponse({"error": "Invalid format authorization"}, status=401)
        event_key = parts[1]
        participant_event = get_participant_event(event_key)
        if not participant_event:
            return JsonResponse({"error": "ParticipantEvent not found"}, status=404)

        now = timezone.now()
        if not _allow_upload_after_block(participant_event, now):
            return JsonResponse({"error": "Monitoring not started"}, status=403)

        s3_key = None

        if "media" in request.FILES:
            file = request.FILES["media"]

            # Subir archivo de audio/video a S3
            upload_result = s3_service.upload_media_fragment(
                file,
                participant_event.id,
                media_type="video",  # Asumimos video por defecto, se puede detectar del archivo
                timestamp=now,
            )

            if upload_result["success"]:
                s3_key = upload_result["key"]
                presigned_url = upload_result.get("presigned_url")
            else:
                return JsonResponse(
                    {
                        "error": f"Failed to upload media fragment: {upload_result['error']}"
                    },
                    status=500,
                )
        else:
            try:
                data = json.loads(request.body or "{}")
            except json.JSONDecodeError:
                data = {}
            s3_key = data.get("s3_key") or request.POST.get("s3_key")
            presigned_url = (
                s3_service.generate_presigned_url(s3_key) if s3_key else None
            )

        if not s3_key:
            return JsonResponse(
                {"error": "Missing media file or s3_key"},
                status=400,
            )

        # Guardar el log con la URL de S3
        ParticipantLog.objects.create(
            name="audio/video",
            url=s3_key,  # guardamos la key en el campo url
            message="Media Capture",
            participant_event=participant_event,
        )
        return JsonResponse(
            {
                "status": "success",
                "s3_key": s3_key,
                "url": presigned_url,
            }
        )

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# Endpoints app web


@csrf_exempt
@jwt_required()
@require_POST
def send_key_emails(request, event_id):
    try:
        # Obtener participantIds del body
        try:
            data = json.loads(request.body.decode("utf-8"))
            participant_ids = data.get("participantIds", [])
        except Exception:
            participant_ids = []

        user = request.user
        try:
            user_role = UserRole.objects.get(user=user)
        except UserRole.DoesNotExist:
            return JsonResponse({"error": "Usuario no encontrado"}, status=404)

        # Verificar que el evento existe
        try:
            event = Event.objects.get(id=event_id)
        except Event.DoesNotExist:
            return JsonResponse({"error": "Evento no encontrado"}, status=404)

        # Permitir solo si es evaluador del evento, admin o superadmin
        if str(event.evaluator_id) != str(user.id) and user_role.role not in [
            "superadmin",
            "admin",
        ]:
            return JsonResponse(
                {
                    "error": "Solo el evaluador asignado o administrador puede enviar correos"
                },
                status=403,
            )

        # Llamar a la función de envío de correos
        result = send_emails(event_id, participant_ids if participant_ids else None)

        if result["success"]:
            return JsonResponse({"success": True, "sent": result["sent"]})
        else:
            return JsonResponse({"error": result["error"]}, status=400)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@csrf_exempt
@jwt_required()
def participants(request):
    """Endpoint para listar y crear participantes"""

    if request.method == "GET":
        # Obtener todos los participantes
        participants = (
            Participant.objects.all().prefetch_related("events").order_by("-created_at")
        )

        # Filtrar por búsqueda si se proporciona un término
        search_term = request.GET.get("search", "")
        if search_term:
            participants = participants.filter(
                Q(name__icontains=search_term) | Q(email__icontains=search_term)
            )

        # Preparar los datos para la respuesta
        participants_data = []
        for participant in participants:
            # Asignar un color según el ID (para la UI)
            colors = ["blue", "green", "purple", "red", "yellow", "indigo", "pink"]
            color = f"bg-{colors[participant.id % len(colors)]}-200"

            # Obtener eventos relacionados
            events_data = [
                {
                    "id": event.id,
                    "name": event.name,
                    "date": (
                        event.start_date.strftime("%d/%m/%Y")
                        if event.start_date
                        else ""
                    ),
                    "status": event.status,
                }
                for event in participant.events.all()
            ]

            participants_data.append(
                {
                    "id": participant.id,
                    "name": participant.name,
                    "email": participant.email,
                    "initials": participant.get_initials(),
                    "color": color,
                    "events": events_data,
                }
            )

        return JsonResponse({"participants": participants_data})

    elif request.method == "POST":
        try:
            # Crear un nuevo participante
            data = json.loads(request.body)

            # Definir variables
            first_name = data.get("first_name", "").strip()
            last_name = data.get("last_name", "").strip()
            email = data.get("email", "").strip()
            full_name = f"{first_name} {last_name}".strip()

            # Validar campos obligatorios
            if not first_name:
                return JsonResponse(
                    {"error": 'El campo "nombre" es obligatorio'}, status=400
                )
            if not last_name:
                return JsonResponse(
                    {"error": 'El campo "apellidos" es obligatorio'}, status=400
                )
            if not email:
                return JsonResponse(
                    {"error": 'El campo "correo" es obligatorio'}, status=400
                )

            # Validar formato de email
            try:
                validate_email(email)
            except ValidationError:
                return JsonResponse({"error": "El correo no es válido"}, status=400)

            # Validar unicidad de email
            if Participant.objects.filter(email__iexact=email).exists():
                return JsonResponse(
                    {"error": "Ya existe un participante con ese correo"}, status=400
                )

            # Crear el participante
            participant = Participant.objects.create(
                first_name=first_name,
                last_name=last_name,
                name=full_name,
                email=email,
            )

            # Responder con éxito
            return JsonResponse(
                {
                    "id": participant.id,
                    "name": participant.name,
                    "message": "Participante creado exitosamente",
                }
            )

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Método no permitido"}, status=405)


@csrf_exempt
@jwt_required()
@require_GET
def export_participants(request):
    """Exporta participantes existentes a Excel con columnas: ID, Nombre, Apellidos, Email."""

    participants = Participant.objects.all().order_by("created_at")
    filename = "participantes.xlsx"

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Participantes"
    ws.append(["ID", "Nombre", "Apellidos", "Email"])

    # Agregar datos existentes
    for p in participants:
        ws.append([p.id, p.first_name, p.last_name, p.email])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@csrf_exempt
@jwt_required()
@require_POST
def import_participants(request):
    """Importa participantes desde Excel con validación previa completa.

    Flujo:
    1. Valida TODAS las filas primero
    2. Si hay errores: retorna solo filas con error (no crea nada)
    3. Si todo válido: crea/actualiza todo en transacción atómica
    """

    # Validar que sea archivo Excel
    if "file" not in request.FILES:
        return JsonResponse({"error": "Se requiere un archivo Excel"}, status=400)

    up_file = request.FILES["file"]

    # Validar tipo de archivo
    filename = getattr(up_file, "name", "")
    content_type = getattr(up_file, "content_type", "")
    if not (
        filename.lower().endswith(".xlsx")
        or content_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        return JsonResponse(
            {"error": "El archivo debe ser un Excel (.xlsx)"}, status=400
        )

    # Leer Excel
    try:
        wb = load_workbook(up_file, data_only=True)
        ws = wb.active
    except Exception as e:
        return JsonResponse(
            {"error": f"No se pudo leer el Excel: {str(e)}"}, status=400
        )

    # Validar cabeceras
    headers = []
    for cell in ws[1]:
        value = (
            (cell.value or "").strip()
            if isinstance(cell.value, str)
            else (cell.value or "")
        )
        headers.append(str(value))

    normalized = [h.strip().lower() for h in headers]
    expected = ["id", "nombre", "apellidos", "email"]
    if normalized[:4] != expected:
        return JsonResponse(
            {
                "error": "Formato inválido. Las primeras 4 columnas deben ser: ID, Nombre, Apellidos, Email",
                "headers": headers,
            },
            status=400,
        )

    # FASE 1: VALIDAR TODAS LAS FILAS
    rows_to_process = []
    seen_emails = set()
    has_errors = False

    # Pre-lectura para obtener todos los IDs presentes en el Excel
    ids_in_excel = set()
    raw_data = []

    for row_idx in range(2, ws.max_row + 1):
        pid = ws.cell(row=row_idx, column=1).value
        fn = ws.cell(row=row_idx, column=2).value
        ln = ws.cell(row=row_idx, column=3).value
        em = ws.cell(row=row_idx, column=4).value

        if not any([pid, fn, ln, em]):
            continue

        raw_data.append({"row_idx": row_idx, "pid": pid, "fn": fn, "ln": ln, "em": em})

        if pid:
            try:
                ids_in_excel.add(int(pid))
            except (ValueError, TypeError):
                pass  # Se validará después

    for item in raw_data:
        row_idx = item["row_idx"]
        pid = item["pid"]
        fn = item["fn"]
        ln = item["ln"]
        em = item["em"]

        # Procesar ID
        participant_id = None
        id_errors = []
        if pid:
            try:
                participant_id = int(pid)
                # Verificar que el ID existe
                if not Participant.objects.filter(id=participant_id).exists():
                    id_errors.append(f"ID {participant_id} no existe en el sistema")
            except (ValueError, TypeError):
                id_errors.append("ID debe ser un número válido")

        # Validar campos
        fn, ln, em, field_errors = _validate_participant_fields(
            fn, ln, em, seen_emails, participant_id, ids_in_excel
        )

        errors = id_errors + field_errors

        rows_to_process.append(
            {
                "row_number": row_idx,
                "id": participant_id,
                "first_name": fn,
                "last_name": ln,
                "email": em,
                "errors": errors,
                "is_update": bool(participant_id),
            }
        )

        if errors:
            has_errors = True

    # Si hay errores: retornar SOLO las filas con error
    if has_errors:
        error_rows = [r for r in rows_to_process if r["errors"]]
        return JsonResponse(
            {
                "success": False,
                "message": "Se encontraron errores. Corrija el archivo y vuelva a importar.",
                "rows": error_rows,
                "total_errors": len(error_rows),
            },
            status=400,
        )

    # FASE 2: CREAR/ACTUALIZAR TODO EN TRANSACCIÓN ATÓMICA
    created_count = 0
    updated_count = 0
    deleted_count = 0

    try:
        with transaction.atomic():
            # 1. Eliminar participantes que no están en el Excel
            # ids_in_excel contiene todos los IDs válidos del archivo
            participants_to_delete = Participant.objects.exclude(id__in=ids_in_excel)
            deleted_count = participants_to_delete.count()

            if deleted_count > 0:
                try:
                    participants_to_delete.delete()
                except RestrictedError:
                    # Identificar qué participantes causaron el error
                    restricted_participants = participants_to_delete.filter(
                        participant_events__isnull=False
                    ).distinct()

                    restricted_names = [
                        f"<strong>{name}</strong>"
                        for name in restricted_participants.values_list(
                            "name", flat=True
                        )
                    ]

                    if len(restricted_names) > 3:
                        names_str = (
                            ", ".join(restricted_names[:3])
                            + f" y {len(restricted_names) - 3} más"
                        )
                    else:
                        names_str = ", ".join(restricted_names)

                    return JsonResponse(
                        {
                            "success": False,
                            "message": f"No se pudieron eliminar los siguientes participantes porque están asociados a uno o más eventos: {names_str}. Desvincúlelos de los eventos antes de proceder.",
                        },
                        status=400,
                    )

            # 2. Actualizar/Crear
            # Para evitar errores de unicidad al intercambiar emails (swap),
            # primero actualizamos todos los emails de los registros a actualizar a un valor temporal.
            # Solo es necesario si hay actualizaciones.

            updates = [r for r in rows_to_process if r["is_update"]]
            creates = [r for r in rows_to_process if not r["is_update"]]

            if updates:
                update_ids = [r["id"] for r in updates]
                # Usamos update() masivo con Case/When o iteramos?
                # Iterar y guardar es lento pero seguro para señales (si las hubiera).
                # Para evitar unique constraint, seteamos emails temporales.
                for r in updates:
                    # Solo necesitamos hacerlo si el email cambia, pero para simplificar lo hacemos a todos
                    # o mejor, solo a aquellos cuyo nuevo email ya existe en BD (aunque sea en otro usuario que también se actualizará).
                    # La estrategia más robusta es setear todos a temporal.
                    p = Participant.objects.get(id=r["id"])
                    p.email = f"temp_{p.id}_{timezone.now().timestamp()}@temp.com"
                    p.save()

            # Ahora aplicamos los valores reales
            for row in updates:
                participant = Participant.objects.get(id=row["id"])
                participant.first_name = row["first_name"]
                participant.last_name = row["last_name"]
                participant.name = f"{row['first_name']} {row['last_name']}".strip()
                participant.email = row["email"]
                participant.save()
                updated_count += 1

                # Actualizar claves de eventos asociados (igual que en edición individual)
                # Al cambiar el email, el hash del event_key cambia
                for pe in ParticipantEvent.objects.filter(participant=participant):
                    pe.save()  # El método save() llama a generate_event_key()

            for row in creates:
                Participant.objects.create(
                    first_name=row["first_name"],
                    last_name=row["last_name"],
                    name=f"{row['first_name']} {row['last_name']}".strip(),
                    email=row["email"],
                )
                created_count += 1

            response_data = {
                "success": True,
                "message": "Importación exitosa",
                "created": created_count,
                "updated": updated_count,
                "deleted": deleted_count,
                "total_processed": len(rows_to_process),
            }

            return JsonResponse(response_data)

    except Exception as e:
        logger.error(f"Error en importación de participantes: {str(e)}")
        return JsonResponse(
            {"error": f"Error al procesar la importación: {str(e)}"},
            status=500,
        )


@csrf_exempt
@jwt_required(roles=["admin", "superadmin"])
def participant_detail(request, participant_id):
    """Endpoint para obtener, actualizar o eliminar un participante específico"""
    try:
        participant = Participant.objects.get(id=participant_id)
    except Participant.DoesNotExist:
        return JsonResponse({"error": "Participante no encontrado"}, status=404)

    if request.method == "GET":
        # Preparar respuesta
        participant_data = {
            "id": participant.id,
            "first_name": participant.first_name,
            "last_name": participant.last_name,
            "email": participant.email,
        }

        return JsonResponse({"participant": participant_data})

    elif request.method == "PUT":
        try:
            # Actualizar el participante
            data = json.loads(request.body)
            changed = False
            email_changed = False

            # Validar campos obligatorios
            first_name = data.get("first_name", "").strip()
            last_name = data.get("last_name", "").strip()
            email = data.get("email", "").strip()
            full_name = f"{first_name} {last_name}".strip()

            if not first_name:
                return JsonResponse(
                    {"error": 'El campo "nombre" es obligatorio'}, status=400
                )
            if not last_name:
                return JsonResponse(
                    {"error": 'El campo "apellidos" es obligatorio'}, status=400
                )
            if not email:
                return JsonResponse(
                    {"error": 'El campo "correo" es obligatorio'}, status=400
                )

            # Validar formato de email
            try:
                validate_email(email)
            except ValidationError:
                return JsonResponse({"error": "El correo no es válido"}, status=400)

            # Validar unicidad de email (excepto el propio participante)
            if (
                Participant.objects.filter(email__iexact=email)
                .exclude(id=participant.id)
                .exists()
            ):
                return JsonResponse(
                    {"error": "Ya existe un participante con ese correo"}, status=400
                )

            # Solo actualizar si cambió el valor
            if participant.first_name != first_name:
                participant.first_name = first_name
                changed = True
            if participant.last_name != last_name:
                participant.last_name = last_name
                changed = True
            if participant.name != full_name:
                participant.name = full_name
                changed = True
            if participant.email != email:
                participant.email = email
                changed = True
                email_changed = True

            if changed:
                participant.save()
                # Si el correo cambió, actualiza los event_key de los eventos asociados
                if email_changed:
                    for pe in ParticipantEvent.objects.filter(participant=participant):
                        pe.generate_event_key()
                        pe.save()

            return JsonResponse(
                {
                    "message": "Participante actualizado exitosamente",
                    "id": participant.id,
                }
            )

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    elif request.method == "DELETE":
        # Eliminar participante
        try:
            participant_name = participant.name
            participant.delete()
            return JsonResponse(
                {"message": f'Participante "{participant_name}" eliminado exitosamente'}
            )
        except Exception as e:
            return JsonResponse(
                {
                    "error": "No se pudo eliminar el participante porque está asociado a uno o más eventos."
                },
                status=400,
            )

    return JsonResponse({"error": "Método no permitido"}, status=405)


@csrf_exempt
@jwt_required()
def events(request):

    user = request.user
    try:
        user_role = UserRole.objects.get(user=user).role
    except UserRole.DoesNotExist:
        user_role = None

    if request.method == "GET":

        # Filtrar eventos según rol: admin/superadmin -> todos, evaluator -> solo sus eventos
        if user_role in ("admin", "superadmin"):
            events = Event.objects.all().order_by("start_date")
        elif user_role == "evaluator":
            events = Event.objects.filter(evaluator__id=user.id).order_by("start_date")
        else:
            # Comportamiento por defecto (sin token o rol desconocido): listar todos
            events = Event.objects.all().order_by("start_date")

        events_data = []
        for event in events:
            participant_count = Participant.objects.filter(events=event).count()

            # Formatear fecha y hora para el frontend
            date_str = event.start_date.strftime("%d/%m/%Y") if event.start_date else ""
            time_str = event.start_date.strftime("%H:%M %p") if event.start_date else ""
            close_date_str = (
                event.close_date.strftime("%d/%m/%Y") if event.close_date else ""
            )
            close_time_str = (
                event.close_date.strftime("%H:%M %p") if event.close_date else ""
            )
            # Duración en minutos
            duration_minutes = event.duration if event.duration else 0
            # Fecha de fin
            end_date_str = event.end_date.strftime("%d/%m/%Y") if event.end_date else ""
            end_time_str = event.end_date.strftime("%H:%M %p") if event.end_date else ""

            # Mapear estado interno a formato de presentación
            status_mapping = {
                "programado": "Programado",
                "en_progreso": "En progreso",
                "completado": "Completado",
            }

            display_status = status_mapping.get(event.status, "Programado")

            # Serializar el evaluador como diccionario en lugar del objeto completo
            evaluator_data = None
            if event.evaluator:
                evaluator_data = (
                    f"{event.evaluator.first_name} {event.evaluator.last_name}"
                )

            events_data.append(
                {
                    "id": event.id,
                    "name": event.name,
                    "startDate": date_str,
                    "startTime": time_str,
                    "closeDate": close_date_str,
                    "closeTime": close_time_str,
                    "duration": duration_minutes,
                    "endDate": end_date_str,
                    "endTime": end_time_str,
                    "participants": participant_count,
                    "status": display_status,
                    "evaluator": evaluator_data,
                }
            )

        return JsonResponse({"events": events_data})

    elif request.method == "POST":

        # Solo admin/superadmin pueden crear eventos
        if user_role not in ("admin", "superadmin"):
            return JsonResponse(
                {"error": "No tienes permisos para crear eventos"},
                status=403,
            )

        # Crear un nuevo evento
        try:
            data = json.loads(request.body)

            # =======================
            # Validaciones
            # =======================
            field_names = {
                "eventName": "nombre del evento",
                "description": "descripción",
                "startDate": "fecha de inicio",
                "startTime": "hora de inicio",
                "closeTime": "fecha de fin",
                "evaluator": "evaluador",
                "duration": "duración",
            }

            # Validar campos
            for field in field_names:
                if not data.get(field):
                    return JsonResponse(
                        {
                            "error": f"El campo {field_names.get(field, field)} es obligatorio"
                        },
                        status=400,
                    )

            # Validar nombre de evento único y longitud mínima
            event_name = data.get("eventName", "").strip()
            if Event.objects.filter(name__iexact=event_name).exists():
                return JsonResponse(
                    {"error": "Ya existe un evento con ese nombre"}, status=400
                )
            if len(event_name) < 4:
                return JsonResponse(
                    {"error": "El nombre del evento debe tener al menos 4 caracteres"},
                    status=400,
                )

            # Validar descripción mínima
            description = data.get("description", "")
            if len(description.strip()) < 5:
                return JsonResponse(
                    {"error": "La descripción debe tener al menos 5 caracteres"},
                    status=400,
                )

            # =======================
            # Validar duración
            # =======================
            duration_minutes = data.get("duration")
            try:
                duration_minutes = int(duration_minutes)
                if duration_minutes < 15:
                    return JsonResponse(
                        {"error": "La duración debe ser de al menos 15 minutos"},
                        status=400,
                    )
                if duration_minutes > 300:  # 5 horas
                    return JsonResponse(
                        {"error": "La duración no puede exceder 5 horas"},
                        status=400,
                    )
            except (ValueError, TypeError):
                return JsonResponse(
                    {"error": "La duración debe ser un número válido de minutos"},
                    status=400,
                )

            # =======================
            # Validar fechas y horas
            # =======================
            tz_str = (data.get("timezone") or "").strip() or "America/Guayaquil"
            try:
                user_tz = ZoneInfo(tz_str)
            except Exception:
                user_tz = ZoneInfo("America/Guayaquil")

            start_date_str = data.get("startDate", "")
            start_time_str = data.get("startTime", "")
            close_time_str = data.get("closeTime", "")

            # Parsear fecha y hora de inicio
            try:
                start_naive = datetime.strptime(
                    f"{start_date_str} {start_time_str}", "%Y-%m-%d %H:%M"
                )
            except ValueError:
                try:
                    start_naive = datetime.strptime(
                        f"{start_date_str} {start_time_str}", "%d/%m/%Y %H:%M"
                    )
                except ValueError:
                    return JsonResponse(
                        {"error": "Formato de fecha/hora de inicio inválido"},
                        status=400,
                    )

            # Parsear hora de cierre (mismo día)
            try:
                close_naive = datetime.strptime(
                    f"{start_date_str} {close_time_str}", "%Y-%m-%d %H:%M"
                )
            except ValueError:
                try:
                    close_naive = datetime.strptime(
                        f"{start_date_str} {close_time_str}", "%d/%m/%Y %H:%M"
                    )
                except ValueError:
                    return JsonResponse(
                        {"error": "Formato de hora de cierre inválido"}, status=400
                    )

            # Convertir a zona horaria del usuario
            start_local = timezone.make_aware(start_naive, user_tz)
            close_local = timezone.make_aware(close_naive, user_tz)

            # Validar que la fecha/hora de inicio sea mayor a la actual en la zona del usuario
            now_local = timezone.localtime(timezone.now(), user_tz).replace(
                second=0, microsecond=0
            )
            if start_local <= now_local:
                return JsonResponse(
                    {
                        "error": "La fecha y hora de inicio deben ser mayor a la fecha y hora actual"
                    },
                    status=400,
                )

            # Validar que la hora de cierre sea al menos 5 minutos después de la hora de inicio
            min_close_local = start_local + timedelta(minutes=5)
            max_close_local = start_local + timedelta(minutes=30)
            if close_local < min_close_local:
                return JsonResponse(
                    {
                        "error": "La hora de cierre debe ser al menos 5 minutos después de la hora de inicio"
                    },
                    status=400,
                )
            if close_local > max_close_local:
                return JsonResponse(
                    {
                        "error": "La hora de cierre no puede ser más de 30 minutos después de la hora de inicio"
                    },
                    status=400,
                )

            # Convertir a UTC para guardar en la BD
            start_utc = start_local.astimezone(ZoneInfo("UTC"))
            close_utc = close_local.astimezone(ZoneInfo("UTC"))

            # Calcular end_date (closeTime + duración en minutos)
            end_utc = close_utc + timedelta(minutes=duration_minutes)

            # =======================
            # 2) Participantes
            # =======================
            participants = data.get("participants", [])
            selected_participants = [
                c for c in participants if c.get("selected", False)
            ]

            # Validar que los participantes no tengan eventos solapados
            conflicting_participants = []
            for participant_data in selected_participants:
                cid = participant_data.get("id")
                if cid:
                    try:
                        participant = Participant.objects.get(id=cid)
                        # Buscar eventos donde el participante esté asignado y se solapen
                        overlapping = Event.objects.filter(
                            participants=participant,
                            start_date__lt=end_utc,
                            end_date__gt=start_utc,
                        )
                        if overlapping.exists():
                            conflicting_participants.append(participant.name)
                    except Participant.DoesNotExist:
                        pass

            if conflicting_participants:
                return JsonResponse(
                    {
                        "error": f"Los siguientes participantes ya tienen eventos en ese rango de fechas: {', '.join(conflicting_participants)}"
                    },
                    status=400,
                )

            # =======================
            # 3) Crear evento
            # =======================

            evaluator_id = data.get("evaluator", "")

            # Obtener la instancia del evaluador
            try:
                evaluator_instance = CustomUser.objects.get(id=evaluator_id)
            except CustomUser.DoesNotExist:
                return JsonResponse(
                    {"error": "El evaluador seleccionado no existe"}, status=400
                )

            # Validar que el evaluador no tenga eventos solapados (por rango de fechas)
            overlapping_events = Event.objects.filter(
                evaluator=evaluator_instance,
                start_date__lt=end_utc,
                end_date__gt=start_utc,
            )
            if overlapping_events.exists():
                return JsonResponse(
                    {
                        "error": "El evaluador ya tiene otro evento en ese rango de fechas"
                    },
                    status=400,
                )

            new_event = Event.objects.create(
                name=data.get("eventName", ""),
                description=data.get("description", ""),
                start_date=start_utc,
                close_date=close_utc,
                duration=duration_minutes,
                end_date=end_utc,
                evaluator=evaluator_instance,
                status="programado",
            )

            # =======================
            # 4) Participantes (asociar a través de ManyToMany)
            # =======================
            for participant_data in selected_participants:
                cid = participant_data.get("id")
                if cid:
                    try:
                        participant = Participant.objects.get(id=cid)
                        ParticipantEvent.objects.get_or_create(
                            event=new_event, participant=participant
                        )
                    except Participant.DoesNotExist:
                        pass

            # Al final, después de crear el evento y asignar participantes:
            blocked_website_ids = data.get("blockedWebsites", [])
            if blocked_website_ids:
                for website_id in blocked_website_ids:
                    try:
                        website = Website.objects.get(id=website_id)
                        BlockedHost.objects.get_or_create(
                            event=new_event, website=website
                        )
                    except Website.DoesNotExist:
                        pass

            return JsonResponse({"success": True, "id": new_event.id})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Método no permitido"}, status=405)


@csrf_exempt
@jwt_required()
def notify_proxy_blocked_hosts_update(request, event_id):
    """Notifica al proxy manager que se actualizaron los hosts bloqueados de un evento
    y devuelve los dominios que necesitan limpieza de caché"""
    if request.method == "POST":
        try:
            cache_key = f"proxy_blocklist_version:{event_id}"
            cache.set(cache_key, int(time.time() * 1000), None)
            # Sistema de señales eliminado - ya no es necesario con arquitectura HTTP directa
            return JsonResponse(
                {
                    "success": True,
                    "message": f"Signal system deprecated - using direct HTTP architecture for event {event_id}",
                }
            )

        except Exception as e:
            logger.error(f"Error notifying proxy of blocked hosts update: {str(e)}")
            return JsonResponse(
                {"error": "Failed to update proxy instances"}, status=500
            )

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@jwt_required()
def event_detail(request, event_id):

    user = request.user
    try:
        user_role = UserRole.objects.get(user=user).role
    except UserRole.DoesNotExist:
        user_role = None

    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return JsonResponse({"error": "Evento no encontrado"}, status=404)

    if request.method == "GET":
        # Obtener detalles del evento
        participants = Participant.objects.filter(events=event)
        participants_data = []

        for participant in participants:
            # Obtener el ParticipantEvent para acceder a is_blocked
            participant_event = ParticipantEvent.objects.filter(
                event=event, participant=participant
            ).first()

            participants_data.append(
                {
                    "id": participant.id,
                    "name": participant.name,
                    "email": participant.email,
                    "initials": participant.get_initials(),
                    "color": f"bg-{['blue', 'green', 'purple', 'red', 'yellow', 'indigo', 'pink'][participant.id % 7]}-200",
                    "is_blocked": (
                        participant_event.is_blocked if participant_event else False
                    ),
                    "is_monitoring": (
                        participant_event.is_monitoring if participant_event else False
                    ),
                }
            )

        # Formatear hora para el frontend
        close_time_str = (
            event.close_date.strftime("%H:%M")
            if getattr(event, "close_date", None)
            else ""
        )
        end_date_str = event.end_date.strftime("%d/%m/%Y") if event.end_date else ""
        end_time_str = event.end_date.strftime("%H:%M") if event.end_date else ""

        # Serializar el evaluador como string + id
        evaluator_name = None
        evaluator_id = None
        if event.evaluator:
            evaluator_name = f"{event.evaluator.first_name} {event.evaluator.last_name}"
            evaluator_id = str(event.evaluator.id)

        # Calcular duración en minutos
        duration_minutes = event.duration

        # Obtener páginas bloqueadas (hostnames)
        blocked_hosts = BlockedHost.objects.filter(event=event).select_related(
            "website"
        )
        blocked_websites = [bh.website.hostname for bh in blocked_hosts]

        event_data = {
            "id": event.id,
            "name": event.name,
            "description": event.description,
            "startDate": (
                event.start_date.strftime("%Y-%m-%d") if event.start_date else ""
            ),
            "startTime": (
                event.start_date.strftime("%H:%M") if event.start_date else ""
            ),
            "closeTime": close_time_str,
            "duration": duration_minutes,
            "evaluator": evaluator_name,
            "evaluatorId": evaluator_id,
            "status": event.status,
            "participants": participants_data,
            "endDate": end_date_str,
            "endTime": end_time_str,
            "blockedWebsites": blocked_websites,
        }

        return JsonResponse({"event": event_data})

    elif request.method == "PUT":
        try:
            data = json.loads(request.body)

            # ========================================
            # VALIDACIÓN DE ESTADO DEL EVENTO
            # ========================================

            # Si el evento está completado, no se puede editar nada
            if event.status == "completado":
                return JsonResponse(
                    {"error": "No se puede editar un evento que ya ha sido completado"},
                    status=400,
                )

            # Si el evento está en progreso, validar que no se cambien fecha/hora de inicio
            if event.status == "en_progreso":
                incoming_start_date = data.get("startDate", "").strip()
                incoming_start_time = data.get("startTime", "").strip()

                # Parsear fecha/hora de inicio entrante
                try:
                    incoming_start_naive = datetime.strptime(
                        f"{incoming_start_date} {incoming_start_time}", "%Y-%m-%d %H:%M"
                    )
                except ValueError:
                    try:
                        incoming_start_naive = datetime.strptime(
                            f"{incoming_start_date} {incoming_start_time}",
                            "%d/%m/%Y %H:%M",
                        )
                    except ValueError:
                        return JsonResponse(
                            {"error": "Formato de fecha/hora de inicio inválido"},
                            status=400,
                        )

                # Obtener timezone del usuario
                tz_str = (data.get("timezone") or "").strip() or "America/Guayaquil"
                try:
                    user_tz = ZoneInfo(tz_str)
                except Exception:
                    user_tz = ZoneInfo("America/Guayaquil")

                incoming_start_local = timezone.make_aware(
                    incoming_start_naive, user_tz
                )
                incoming_start_utc = incoming_start_local.astimezone(ZoneInfo("UTC"))

                # Comparar con la fecha/hora de inicio actual (sin tolerancia)
                if event.start_date != incoming_start_utc:
                    return JsonResponse(
                        {
                            "error": "No se puede modificar la fecha u hora de inicio de un evento en progreso"
                        },
                        status=400,
                    )

            # ========================================
            # VALIDACIONES GENERALES (continúan igual)
            # ========================================

            # Validaciones requeridas
            field_names = {
                "eventName": "nombre del evento",
                "description": "descripción",
                "startDate": "fecha de inicio",
                "startTime": "hora de inicio",
                "closeTime": "hora de cierre",
                "evaluator": "evaluador",
                "duration": "duración",
            }
            # Si es un evaluador, no exigir el campo "evaluator" en el payload
            if user_role == "evaluator":
                field_names.pop("evaluator", None)

            for field in field_names:
                if not data.get(field):
                    return JsonResponse(
                        {
                            "error": f"El campo {field_names.get(field, field)} es obligatorio"
                        },
                        status=400,
                    )

            # Validar nombre de evento único y longitud mínima
            event_name = data.get("eventName", "").strip()
            if (
                Event.objects.filter(name__iexact=event_name)
                .exclude(id=event.id)
                .exists()
            ):
                return JsonResponse(
                    {"error": "Ya existe un evento con ese nombre"}, status=400
                )
            if len(event_name) < 4:
                return JsonResponse(
                    {"error": "El nombre del evento debe tener al menos 4 caracteres"},
                    status=400,
                )

            # Validar descripción mínima
            description = data.get("description", "")
            if len(description.strip()) < 5:
                return JsonResponse(
                    {"error": "La descripción debe tener al menos 5 caracteres"},
                    status=400,
                )

            # =======================
            # Validar duración
            # =======================
            duration_minutes = data.get("duration")
            try:
                duration_minutes = int(duration_minutes)
                if duration_minutes < 15:
                    return JsonResponse(
                        {"error": "La duración debe ser de al menos 15 minutos"},
                        status=400,
                    )
                if duration_minutes > 300:  # 5 horas
                    return JsonResponse(
                        {"error": "La duración no puede exceder 5 horas"},
                        status=400,
                    )
            except (ValueError, TypeError):
                return JsonResponse(
                    {"error": "La duración debe ser un número válido de minutos"},
                    status=400,
                )

            # =======================
            # Validar fechas y horas
            # =======================
            tz_str = (data.get("timezone") or "").strip() or "America/Guayaquil"
            try:
                user_tz = ZoneInfo(tz_str)
            except Exception:
                user_tz = ZoneInfo("America/Guayaquil")

            start_date_str = data.get("startDate", "")
            start_time_str = data.get("startTime", "")
            close_time_str = data.get("closeTime", "")

            # Parsear fecha y hora de inicio
            try:
                start_naive = datetime.strptime(
                    f"{start_date_str} {start_time_str}", "%Y-%m-%d %H:%M"
                )
            except ValueError:
                try:
                    start_naive = datetime.strptime(
                        f"{start_date_str} {start_time_str}", "%d/%m/%Y %H:%M"
                    )
                except ValueError:
                    return JsonResponse(
                        {"error": "Formato de fecha/hora de inicio inválido"},
                        status=400,
                    )

            # Parsear hora de cierre (mismo día)
            try:
                close_naive = datetime.strptime(
                    f"{start_date_str} {close_time_str}", "%Y-%m-%d %H:%M"
                )
            except ValueError:
                try:
                    close_naive = datetime.strptime(
                        f"{start_date_str} {close_time_str}", "%d/%m/%Y %H:%M"
                    )
                except ValueError:
                    return JsonResponse(
                        {"error": "Formato de hora de cierre inválido"}, status=400
                    )

            # Convertir a zona horaria del usuario
            start_local = timezone.make_aware(start_naive, user_tz)
            close_local = timezone.make_aware(close_naive, user_tz)

            # Validar que la fecha/hora de inicio sea posterior a ahora
            # OMITIR esta validación si el evento está en progreso (no se permite cambiar inicio, pero sí otros campos)
            if event.status != "en_progreso":
                now_local = timezone.localtime(timezone.now(), user_tz).replace(
                    second=0, microsecond=0
                )
                if start_local <= now_local:
                    return JsonResponse(
                        {
                            "error": "La fecha y hora de inicio deben ser mayor a la fecha y hora actual"
                        },
                        status=400,
                    )

            # Validar que la hora de cierre sea al menos 5 minutos después de la hora de inicio
            min_close_local = start_local + timedelta(minutes=5)
            max_close_local = start_local + timedelta(minutes=30)
            if close_local < min_close_local:
                return JsonResponse(
                    {
                        "error": "La hora de cierre debe ser al menos 5 minutos después de la hora de inicio"
                    },
                    status=400,
                )
            if close_local > max_close_local:
                return JsonResponse(
                    {
                        "error": "La hora de cierre no puede ser más de 30 minutos después de la hora de inicio"
                    },
                    status=400,
                )

            # Convertir a UTC para guardar en la BD
            start_utc = start_local.astimezone(ZoneInfo("UTC"))
            close_utc = close_local.astimezone(ZoneInfo("UTC"))
            end_utc = close_utc + timedelta(minutes=duration_minutes)

            # Evaluador
            evaluator_id = data.get("evaluator", "")
            if user_role == "evaluator":
                # Un evaluador no puede cambiar el evaluador asignado
                if evaluator_id and str(evaluator_id) != str(event.evaluator_id):
                    return JsonResponse(
                        {
                            "error": "No tienes permiso para cambiar el evaluador del evento"
                        },
                        status=403,
                    )
                evaluator_instance = event.evaluator
            else:
                try:
                    evaluator_instance = CustomUser.objects.get(id=evaluator_id)
                except CustomUser.DoesNotExist:
                    return JsonResponse(
                        {"error": "El evaluador seleccionado no existe"}, status=400
                    )

            # Validar que el evaluador no tenga eventos solapados (excluyendo el actual)
            overlapping_events = Event.objects.filter(
                evaluator=evaluator_instance,
                start_date__lt=end_utc,
                end_date__gt=start_utc,
            ).exclude(id=event.id)
            if overlapping_events.exists():
                return JsonResponse(
                    {
                        "error": "El evaluador ya tiene otro evento en ese rango de fechas"
                    },
                    status=400,
                )

            # Validar que los participantes no tengan eventos solapados (excluyendo el evento actual)
            participants = data.get("participants", [])
            selected_participants = [
                c for c in participants if c.get("selected", False)
            ]

            conflicting_participants = []
            for participant_data in selected_participants:
                cid = participant_data.get("id")
                if cid:
                    try:
                        participant = Participant.objects.get(id=cid)
                        overlapping = Event.objects.filter(
                            participants=participant,
                            start_date__lt=end_utc,
                            end_date__gt=start_utc,
                        ).exclude(id=event.id)
                        if overlapping.exists():
                            conflicting_participants.append(participant.name)
                    except Participant.DoesNotExist:
                        pass

            if conflicting_participants:
                return JsonResponse(
                    {
                        "error": f"Los siguientes participantes ya tienen eventos en ese rango de fechas: {', '.join(conflicting_participants)}"
                    },
                    status=400,
                )

            # Actualización de campos
            changed = False
            if event.name != event_name:
                event.name = event_name
                changed = True
            if event.description != description:
                event.description = description
                changed = True
            if event.start_date != start_utc:
                event.start_date = start_utc
                changed = True
            if getattr(event, "close_date", None) != close_utc:
                event.close_date = close_utc
                changed = True
            # El evaluador solo puede ser cambiado por admin/superadmin
            if event.evaluator != evaluator_instance:
                event.evaluator = evaluator_instance
                changed = True
            if event.duration != duration_minutes:
                event.duration = duration_minutes
                changed = True
            if event.end_date != end_utc:
                event.end_date = end_utc
                changed = True

            if changed:
                event.save()

            # Participantes
            if "participants" in data:
                try:
                    handle_event_participants(event, data.get("participants", []))
                except Exception as participant_error:
                    print(f"Error processing participantes: {participant_error}")

            # Websites bloqueados
            if "blockedWebsites" in data:
                incoming_ids = set()
                for wid in data.get("blockedWebsites") or []:
                    try:
                        incoming_ids.add(int(wid))
                    except (TypeError, ValueError):
                        continue

                existing_bh = BlockedHost.objects.filter(event=event)
                existing_ids = set(existing_bh.values_list("website_id", flat=True))

                to_add = incoming_ids - existing_ids
                to_remove = existing_ids - incoming_ids

                if to_add:
                    for wid in to_add:
                        try:
                            website = Website.objects.get(id=wid)
                            BlockedHost.objects.get_or_create(
                                event=event, website=website
                            )
                        except Website.DoesNotExist:
                            continue

                if to_remove:
                    BlockedHost.objects.filter(
                        event=event, website_id__in=list(to_remove)
                    ).delete()

                # Sistema de notificación eliminado - ya no necesario con HTTP directo
                if to_add or to_remove:
                    print(
                        f"🔄 | Blocked hosts updated for event {event_id}: +{len(to_add)} -{len(to_remove)}"
                    )

            return JsonResponse({"success": True, "updated": changed})

        except Exception as e:
            print(f"Error updating event: {e}")
            return JsonResponse({"error": str(e)}, status=400)

    elif request.method == "DELETE":

        # Solo admin/superadmin pueden eliminar eventos
        if user_role not in ("admin", "superadmin"):
            return JsonResponse(
                {"error": "No tienes permisos para eliminar eventos"},
                status=403,
            )

        # No permitir eliminar si el evento está en progreso
        if event.status == "en_progreso":
            return JsonResponse(
                {"error": "No se puede eliminar un evento que está en progreso"},
                status=400,
            )

        # Eliminar un evento
        try:
            s3_enabled = s3_service.is_configured()
            media_keys = _collect_event_media_keys(event.id) if s3_enabled else set()
            event_id = event.id

            event.delete()

            if s3_enabled and media_keys:
                try:
                    delete_event_media_from_s3.delay(event_id, sorted(media_keys))
                except Exception as task_error:
                    logger.warning(
                        "Failed to enqueue S3 cleanup for event %s: %s",
                        event_id,
                        task_error,
                    )
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Método no permitido"}, status=405)


@csrf_exempt
@jwt_required()
def websites(request):
    """Endpoint para listar y crear sitios web (páginas bloqueables)"""

    if request.method == "GET":
        sites = Website.objects.all().order_by("hostname")
        sites_data = [
            {
                "id": str(site.id),
                "hostname": site.hostname,
            }
            for site in sites
        ]
        return JsonResponse({"websites": sites_data})

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            hostname = data.get("hostname", "").strip().lower()

            if not hostname:
                return JsonResponse(
                    {"error": "El nombre del sitio es obligatorio"}, status=400
                )

            if not is_valid_domain(hostname):
                return JsonResponse(
                    {"error": "El nombre del sitio no es un dominio válido"}, status=400
                )

            if Website.objects.filter(hostname__iexact=hostname).exists():
                return JsonResponse({"error": "Este sitio web ya existe"}, status=400)

            website = Website.objects.create(hostname=hostname)

            return JsonResponse(
                {
                    "id": str(website.id),
                    "hostname": website.hostname,
                }
            )
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Método no permitido"}, status=405)


@csrf_exempt
@jwt_required()
def website_detail(request, website_id):
    """Endpoint para actualizar o eliminar un sitio web"""

    try:
        website = Website.objects.get(id=website_id)
    except Website.DoesNotExist:
        return JsonResponse({"error": "Sitio web no encontrado"}, status=404)

    if request.method == "PUT":
        try:
            data = json.loads(request.body)
            hostname = data.get("hostname", "").strip().lower()

            if not hostname:
                return JsonResponse(
                    {"error": "El nombre del sitio es obligatorio"}, status=400
                )

            if not is_valid_domain(hostname):
                return JsonResponse(
                    {"error": "El nombre del sitio no es un dominio válido"}, status=400
                )

            if (
                Website.objects.filter(hostname__iexact=hostname)
                .exclude(id=website_id)
                .exists()
            ):
                return JsonResponse({"error": "Este sitio web ya existe"}, status=400)

            website.hostname = hostname
            website.save()

            return JsonResponse(
                {
                    "id": str(website.id),
                    "hostname": website.hostname,
                }
            )
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    elif request.method == "DELETE":
        try:
            website.delete()
            return JsonResponse({"message": "Sitio web eliminado exitosamente"})
        except RestrictedError as e:
            return JsonResponse(
                {"error": "No se puede eliminar porque esta asociado a un evento"},
                status=400,
            )
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Método no permitido"}, status=405)


@csrf_exempt
@jwt_required()
@require_GET
def event_blocked_hosts(request, event_id):
    """Endpoint para obtener los hosts bloqueados de un evento"""

    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return JsonResponse({"error": "Evento no encontrado"}, status=404)

    blocked_hosts = BlockedHost.objects.filter(event=event).select_related("website")
    blocked_website_ids = [str(bh.website.id) for bh in blocked_hosts]

    return JsonResponse({"blocked_website_ids": blocked_website_ids})


@csrf_exempt
@jwt_required()
@require_GET
def evaluaciones(request):
    """
    Endpoint para listar evaluaciones (eventos en progreso y completados).
    Admin/superadmin ven todas, evaluadores solo las suyas.
    """
    user = request.user
    try:
        user_role = UserRole.objects.get(user=user).role
    except UserRole.DoesNotExist:
        user_role = None

    # Filtrar eventos por estado
    estados = ["en_progreso", "completado"]
    if user_role in ("admin", "superadmin"):
        eventos = Event.objects.filter(status__in=estados).order_by("start_date")
    elif user_role == "evaluator":
        eventos = Event.objects.filter(
            status__in=estados, evaluator__id=user.id
        ).order_by("start_date")
    else:
        eventos = Event.objects.filter(status__in=estados).order_by("start_date")

    eventos_data = []
    for evento in eventos:
        participant_count = Participant.objects.filter(events=evento).count()
        eventos_data.append(
            {
                "id": evento.id,
                "name": evento.name,
                "startDate": (
                    evento.start_date.strftime("%d/%m/%Y") if evento.start_date else ""
                ),
                "startTime": (
                    evento.start_date.strftime("%H:%M %p") if evento.start_date else ""
                ),
                "closeDate": (
                    evento.close_date.strftime("%d/%m/%Y") if evento.close_date else ""
                ),
                "closeTime": (
                    evento.close_date.strftime("%H:%M %p") if evento.close_date else ""
                ),
                "duration": evento.duration,
                "endDate": (
                    evento.end_date.strftime("%d/%m/%Y") if evento.end_date else ""
                ),
                "endTime": (
                    evento.end_date.strftime("%H:%M %p") if evento.end_date else ""
                ),
                "participants": participant_count,
                "status": evento.status,
            }
        )

    return JsonResponse({"evaluaciones": eventos_data})


@csrf_exempt
@jwt_required()
@require_GET
def evaluation_detail(request, evaluation_id):
    """
    Endpoint para obtener el detalle de una evaluación (evento).
    Devuelve los datos en el formato de EvaluationDetail.
    """
    try:
        event = Event.objects.get(id=evaluation_id)
    except Event.DoesNotExist:
        return JsonResponse({"error": "Evaluación no encontrada"}, status=404)

    participants = Participant.objects.filter(events=event)
    participants_data = []

    for p in participants:
        # Obtener estado de monitoreo y bloqueo
        monitoring_status = False
        is_blocked_status = False

        try:
            participant_event = ParticipantEvent.objects.get(event=event, participant=p)
            monitoring_status = getattr(participant_event, "is_monitoring", False)
            is_blocked_status = getattr(participant_event, "is_blocked", False)
        except ParticipantEvent.DoesNotExist:
            pass

        participants_data.append(
            {
                "id": str(p.id),
                "name": p.name,
                "initials": p.get_initials() if hasattr(p, "get_initials") else "",
                "monitoring_is_active": monitoring_status,
                "is_blocked": is_blocked_status,
                "color": f"bg-{['blue', 'green', 'purple', 'red', 'yellow', 'indigo', 'pink'][p.id % 7]}-200",
            }
        )

    evaluator_name = None
    if event.evaluator:
        evaluator_name = f"{event.evaluator.first_name} {event.evaluator.last_name}"

    detail = {
        "id": str(event.id),
        "name": event.name,
        "description": event.description,
        "startDate": event.start_date.strftime("%Y-%m-%d") if event.start_date else "",
        "startTime": event.start_date.strftime("%H:%M") if event.start_date else "",
        "closeDate": event.close_date.strftime("%Y-%m-%d") if event.close_date else "",
        "closeTime": event.close_date.strftime("%H:%M") if event.close_date else "",
        "duration": event.duration,
        "endDate": event.end_date.strftime("%Y-%m-%d") if event.end_date else "",
        "endTime": event.end_date.strftime("%H:%M") if event.end_date else "",
        "status": event.status,
        "participants": participants_data,
        "evaluator": evaluator_name,
    }

    return JsonResponse({"event": detail})


# Logs


@csrf_exempt
@jwt_required()
@require_GET
def event_participant_logs(request, event_id, participant_id):
    """Endpoint para obtener todos los logs de un participante específico de un evento"""
    try:
        event = Event.objects.get(id=event_id)
        participant = Participant.objects.get(id=participant_id)

        # Obtener el ParticipantEvent específico
        participant_event = ParticipantEvent.objects.filter(
            event=event, participant=participant
        ).first()

        if not participant_event:
            return JsonResponse(
                {"error": "Participant not found in this event"}, status=404
            )

        # Filtrar logs solo del participante específico
        logs = ParticipantLog.objects.filter(participant_event=participant_event)

        logs_data = []
        for log in logs:
            # Formatear timestamp como string legible
            created_at_formatted = ""
            if hasattr(log, "timestamp") and getattr(log, "timestamp"):
                try:
                    created_at_formatted = log.timestamp.strftime("%d/%m/%Y %H:%M:%S")
                except Exception:
                    created_at_formatted = "N/A"
            else:
                created_at_formatted = "N/A"

            presigned_url = None
            if log.url and s3_service.is_configured():
                presigned_url = s3_service.generate_presigned_url(log.url)

            log_data = {
                "id": log.id,
                "name": log.name,
                "message": log.message,
                "created_at": created_at_formatted,
                "has_file": bool(log.url),
                "file_url": presigned_url,
                "s3_key": log.url,
            }
            logs_data.append(log_data)

        return JsonResponse(
            {
                "event": {"id": event.id, "name": event.name},
                "logs": logs_data,
                "total": len(logs_data),
            }
        )

    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)
    except Participant.DoesNotExist:
        return JsonResponse({"error": "Participant not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@jwt_required()
def participant_connection_stats(request, event_id, participant_id):
    """Endpoint para obtener estadísticas de conexión de un participante en un evento específico"""

    try:
        event = Event.objects.get(id=event_id)
        participant = Participant.objects.get(id=participant_id)

        # Buscar el ParticipantEvent específico del evento y participante
        participant_event = ParticipantEvent.objects.filter(
            event=event, participant=participant
        ).first()

        connection_data = {
            "participant": {
                "id": participant.id,
                "name": participant.name,
                "email": participant.email,
            },
            "total_time_seconds": 0,
            "monitoring_is_active": False,
            "monitoring_last_change": None,
            "monitoring_sessions_count": 0,
        }

        if participant_event:
            # Todos los datos desde ParticipantEvent del evento específico
            connection_data.update(
                {
                    "monitoring_is_active": participant_event.is_monitoring,
                    "total_time_seconds": participant_event.get_total_monitoring_seconds(),
                    "monitoring_last_change": participant_event.monitoring_last_change,
                    "monitoring_sessions_count": participant_event.monitoring_sessions_count,
                }
            )

        return JsonResponse(connection_data)

    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)
    except Participant.DoesNotExist:
        return JsonResponse({"error": "Participant not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@jwt_required()
@require_POST
def cleanup_stale_monitoring_by_logs(request):
    """Stops monitoring sessions when no logs are received for a while."""
    try:
        payload = {}
        if request.body:
            try:
                payload = json.loads(request.body.decode("utf-8"))
            except json.JSONDecodeError:
                payload = {}

        threshold_seconds = payload.get("threshold_seconds", 180)
        try:
            threshold_seconds = int(threshold_seconds)
        except (TypeError, ValueError):
            threshold_seconds = 180

        if threshold_seconds <= 0:
            threshold_seconds = 180

        now = timezone.now()
        cutoff = now - timedelta(seconds=threshold_seconds)

        base_qs = (
            ParticipantEvent.objects.filter(is_monitoring=True)
            .annotate(last_log=Max("participantlog__timestamp"))
            .filter(
                Q(monitoring_current_session_time__lte=cutoff)
                | Q(monitoring_current_session_time__isnull=True)
            )
        )

        stale_qs = base_qs.filter(Q(last_log__lt=cutoff) | Q(last_log__isnull=True))

        results = []
        with transaction.atomic():
            for pe in stale_qs:
                session_seconds = 0
                inactive_seconds = 0
                if pe.monitoring_current_session_time:
                    start_time = pe.monitoring_current_session_time
                    # Si hay último log, calcular hasta ese momento
                    # Si no hay logs, calcular hasta el cutoff (no contar tiempo sin actividad)
                    if pe.last_log and pe.last_log >= start_time:
                        end_time = pe.last_log
                    else:
                        # Sin logs recientes, contar solo hasta el cutoff
                        end_time = cutoff if cutoff >= start_time else start_time
                    
                    if end_time >= start_time:
                        session_seconds = int((end_time - start_time).total_seconds())
                        pe.monitoring_total_duration += session_seconds
                        
                        # Calcular tiempo inactivo (desde último log o cutoff hasta ahora)
                        inactive_seconds = int((now - end_time).total_seconds())

                pe.monitoring_current_session_time = None
                pe.is_monitoring = False
                pe.monitoring_last_change = now

                pe.save(
                    update_fields=[
                        "is_monitoring",
                        "monitoring_total_duration",
                        "monitoring_current_session_time",
                        "monitoring_last_change",
                    ]
                )

                if pe.event_key:
                    cache.delete(f"verify_event_key_{pe.event_key}")

                results.append(
                    {
                        "participant_event_id": pe.id,
                        "event_id": pe.event_id,
                        "participant_id": pe.participant_id,
                        "last_log": pe.last_log.isoformat() if pe.last_log else None,
                        "session_seconds": session_seconds,
                        "inactive_seconds": inactive_seconds,
                    }
                )

        return JsonResponse(
            {
                "success": True,
                "threshold_seconds": threshold_seconds,
                "stale_count": len(results),
                "results": results,
            }
        )
    except Exception as e:
        logger.error(f"Error cleaning stale monitoring by logs: {e}")
        return JsonResponse({"error": "Internal server error"}, status=500)


# =============================
# Endpoints de estado de eventos
# =============================


@csrf_exempt
@require_GET
def pending_start_events(request):
    """Eventos en estado 'programado' cuyo start_date ya pasó o es ahora.

    Se usan para decidir qué eventos arrancar.
    """
    now = timezone.now()
    qs = Event.objects.filter(status="programado", start_date__lte=now).order_by(
        "start_date"
    )
    data = [
        {
            "id": e.id,
            "start_date": e.start_date.isoformat() if e.start_date else None,
        }
        for e in qs
    ]
    return JsonResponse({"results": data})


@csrf_exempt
@require_GET
def pending_finish_events(request):
    """Eventos en estado 'en_progreso' cuyo end_date ya pasó o es ahora.

    Sirve para decidir cuáles finalizar.
    """
    now = timezone.now()
    qs = Event.objects.filter(status="en_progreso", end_date__lte=now).order_by(
        "end_date"
    )
    data = [
        {
            "id": e.id,
            "end_date": e.end_date.isoformat() if e.end_date else None,
        }
        for e in qs
    ]
    return JsonResponse({"results": data})


@csrf_exempt
@jwt_required(roles=["superadmin"])
@require_GET
def expired_events(request):
    """Eventos cuyo start_date (solo fecha) es >= 182 dias antes de hoy (UTC)."""
    today_date = timezone.now().date()
    cutoff_date = today_date - timedelta(days=EVENT_EXPIRATION_DAYS)
    qs = Event.objects.filter(start_date__date__lte=cutoff_date).order_by("start_date")
    data = [
        {
            "id": e.id,
            "start_date": e.start_date.date().isoformat() if e.start_date else None,
        }
        for e in qs
    ]
    return JsonResponse(
        {
            "results": data,
            "cutoff_date": cutoff_date.isoformat(),
            "days": EVENT_EXPIRATION_DAYS,
        }
    )


# Eventos: actualizar estados (start/finish)
@csrf_exempt
@require_POST
def start_event(request, event_id):
    """
    Cambia el estado de un evento de 'programado' a 'en_progreso'.
    """
    try:
        try:
            event = Event.objects.get(id=event_id)
        except Event.DoesNotExist:
            return JsonResponse({"error": "Evento no encontrado"}, status=404)

        if event.status != "programado":
            return JsonResponse(
                {"error": "El evento no está en estado 'programado'"}, status=400
            )

        event.status = "en_progreso"
        event.save()
        return JsonResponse({"success": True, "status": event.status})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@csrf_exempt
@require_POST
def finish_event(request, event_id):
    """
    Cambia el estado de un evento de 'en_progreso' a 'completado'.
    """
    try:
        try:
            event = Event.objects.get(id=event_id)
        except Event.DoesNotExist:
            return JsonResponse({"error": "Evento no encontrado"}, status=404)

        if event.status != "en_progreso":
            return JsonResponse(
                {"error": "El evento no está en estado 'en_progreso'"}, status=400
            )

        event.status = "completado"
        event.save()
        return JsonResponse({"success": True, "status": event.status})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@csrf_exempt
@jwt_required()
def participant_media_files(request, event_id, participant_id):
    """
    Endpoint para listar archivos multimedia de un participante en S3.
    """
    if request.method == "GET":
        try:
            # Verificar que el evento y participante existen
            event = Event.objects.get(id=event_id)
            participant = Participant.objects.get(id=participant_id)

            # Obtener el ParticipantEvent
            participant_event = ParticipantEvent.objects.filter(
                event=event, participant=participant
            ).first()

            if not participant_event:
                return JsonResponse(
                    {"error": "Participant not enrolled in this event"}, status=404
                )

            # Verificar si S3 está configurado
            if not s3_service.is_configured():
                return JsonResponse({"error": "S3 service not configured"}, status=503)

            # Obtener parámetros de filtro
            media_type = request.GET.get("media_type")  # video, audio, screen
            start_date = request.GET.get("start_date")
            end_date = request.GET.get("end_date")

            # Parsear fechas si se proporcionan
            start_datetime = None
            end_datetime = None
            if start_date:
                try:
                    start_datetime = datetime.fromisoformat(start_date)
                except ValueError:
                    return JsonResponse(
                        {"error": "Invalid start_date format"}, status=400
                    )
            if end_date:
                try:
                    end_datetime = datetime.fromisoformat(end_date)
                except ValueError:
                    return JsonResponse(
                        {"error": "Invalid end_date format"}, status=400
                    )

            # Listar archivos multimedia del participante
            media_files = s3_service.list_participant_media(
                participant_event.id,
                media_type=media_type,
                start_date=start_datetime,
                end_date=end_datetime,
            )

            # Organizar archivos por tipo
            files_by_type = {"video": [], "audio": [], "screen": [], "unknown": []}

            total_size = 0
            for file_info in media_files:
                file_type = file_info.get("media_type", "unknown")
                if file_type in files_by_type:
                    files_by_type[file_type].append(file_info)
                else:
                    files_by_type["unknown"].append(file_info)
                total_size += file_info.get("size", 0)

            return JsonResponse(
                {
                    "success": True,
                    "participant": {
                        "id": participant.id,
                        "name": participant.name,
                        "email": participant.email,
                    },
                    "event": {"id": event.id, "name": event.name},
                    "files_by_type": files_by_type,
                    "summary": {
                        "total_files": len(media_files),
                        "total_size_bytes": total_size,
                        "video_files": len(files_by_type["video"]),
                        "audio_files": len(files_by_type["audio"]),
                        "screen_files": len(files_by_type["screen"]),
                    },
                }
            )

        except Event.DoesNotExist:
            return JsonResponse({"error": "Event not found"}, status=404)
        except Participant.DoesNotExist:
            return JsonResponse({"error": "Participant not found"}, status=404)
        except Exception as e:
            logger.error(
                f"Error listing media files for participant {participant_id} in event {event_id}: {e}"
            )
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@jwt_required()
def create_s3_bucket(request):
    """
    Endpoint para crear el bucket de S3 si no existe.
    Solo para administradores.
    """
    if request.method == "POST":
        try:
            # Verificar permisos de administrador
            user_data = get_user_data(request)
            if not user_data or user_data.get("role") != UserRole.ADMIN.value:
                return JsonResponse({"error": "Admin privileges required"}, status=403)

            if not s3_service.is_configured():
                return JsonResponse(
                    {
                        "error": "S3 service not configured. Please check AWS credentials and settings."
                    },
                    status=503,
                )

            # Intentar crear el bucket
            success = s3_service.create_bucket_if_not_exists()

            if success:
                return JsonResponse(
                    {
                        "success": True,
                        "message": f"Bucket '{s3_service.bucket_name}' is ready for use.",
                        "bucket_name": s3_service.bucket_name,
                        "region": s3_service.region,
                    }
                )
            else:
                return JsonResponse(
                    {
                        "error": "Failed to create or access S3 bucket. Check AWS credentials and permissions."
                    },
                    status=500,
                )

        except Exception as e:
            logger.error(f"Error creating S3 bucket: {e}")
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@jwt_required(roles=["admin", "superadmin"])
@require_POST
def block_participants(request, event_id):
    """Bloquea participantes seleccionados de un evento"""
    
    try:
        event = Event.objects.get(id=event_id)
        data = json.loads(request.body)
        participant_ids = data.get("participant_ids", [])

        if not participant_ids:
            return JsonResponse(
                {"error": "No se especificaron participantes"}, status=400
            )

        # Bloquear los participantes seleccionados
        # IMPORTANTE: Guardar el tiempo de monitoreo antes de bloquear
        blocked_count = 0
        now = timezone.now()
        
        with transaction.atomic():
            participant_events = ParticipantEvent.objects.select_for_update().filter(
                event=event, participant_id__in=participant_ids
            )
            
            for pe in participant_events:
                # Si está monitoreando activamente, guardar el tiempo de la sesión actual
                if pe.is_monitoring and pe.monitoring_current_session_time:
                    start_time = pe.monitoring_current_session_time
                    if now >= start_time:
                        session_seconds = int((now - start_time).total_seconds())
                        pe.monitoring_total_duration += session_seconds
                    
                    # Limpiar la sesión actual
                    pe.monitoring_current_session_time = None
                
                # Bloquear el participante
                pe.is_blocked = True
                pe.is_monitoring = False
                pe.monitoring_last_change = now
                
                pe.save(update_fields=[
                    "is_blocked",
                    "is_monitoring", 
                    "monitoring_total_duration",
                    "monitoring_current_session_time",
                    "monitoring_last_change"
                ])
                blocked_count += 1

        return JsonResponse(
            {
                "success": True,
                "message": f"{blocked_count} participante(s) bloqueado(s) exitosamente",
            }
        )

    except Event.DoesNotExist:
        return JsonResponse({"error": "Evento no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ============================================
# ENDPOINTS DE CONSENTIMIENTO INFORMADO
# ============================================


@csrf_exempt
@require_POST
def register_event_consent(request):
    """
    Registra el consentimiento informado de un participante para un evento específico.
    Cumple con la Ley Orgánica de Protección de Datos Personales del Ecuador.

    Header requerido:
        Authorization: Bearer {event_key}

    Body JSON:
        {
            "accepted": true,
            "consent_version": "v1.0"
        }

    Returns:
        200: Consentimiento registrado exitosamente
        400: Datos inválidos o consentimiento no aceptado
        401: Authorization header inválido
        404: Event_key no encontrado
        409: Consentimiento ya existe
        500: Error del servidor
    """
    try:
        # Validar authorization header
        authorization = request.headers.get("Authorization", "")
        if not authorization.startswith("Bearer "):
            return JsonResponse(
                {"error": "Authorization header inválido. Formato: Bearer {event_key}"},
                status=401,
            )

        parts = authorization.split(" ", 1)
        if len(parts) != 2 or not parts[1]:
            return JsonResponse(
                {"error": "Event_key no proporcionado en Authorization header"},
                status=401,
            )

        event_key = parts[1]

        # Obtener participante y evento asociados al event_key
        try:
            participant_event = ParticipantEvent.objects.select_related(
                "participant", "event"
            ).get(event_key=event_key)
            participant = participant_event.participant
            event = participant_event.event
        except ParticipantEvent.DoesNotExist:
            return JsonResponse({"error": "Event_key inválido o no existe"}, status=404)

        # Parsear body JSON
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Body JSON inválido"}, status=400)

        # Validar que el usuario aceptó explícitamente
        accepted = data.get("accepted")
        if accepted is not True:
            return JsonResponse(
                {
                    "error": "Debe aceptar explícitamente el consentimiento informado",
                    "message": "El consentimiento debe ser explícito para continuar con la evaluación",
                },
                status=400,
            )

        # Obtener versión del consentimiento (por defecto v1.0)
        consent_version = data.get("consent_version", "v1.0")

        # Verificar si ya existe consentimiento
        existing_consent = EventConsent.objects.filter(
            participant=participant, event=event
        ).first()

        if existing_consent:
            return JsonResponse(
                {
                    "error": "Ya existe un consentimiento registrado para este evento",
                    "consent": {
                        "accepted_at": existing_consent.accepted_at.isoformat(),
                        "consent_version": existing_consent.consent_version,
                    },
                },
                status=409,
            )

        # Obtener información adicional de la request
        def get_client_ip(request):
            """Obtiene la IP real del cliente considerando proxies"""
            x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
            if x_forwarded_for:
                ip = x_forwarded_for.split(",")[0].strip()
            else:
                ip = request.META.get("REMOTE_ADDR")
            return ip

        ip_address = get_client_ip(request)
        user_agent = request.META.get("HTTP_USER_AGENT", "")

        # Crear registro de consentimiento con transacción atómica
        with transaction.atomic():
            consent = EventConsent.objects.create(
                participant=participant,
                event=event,
                consent_version=consent_version,
                ip_address=ip_address,
                user_agent=user_agent,
            )

            logger.info(
                f"Consentimiento registrado: Participante={participant.email}, "
                f"Evento={event.name}, Versión={consent_version}, IP={ip_address}"
            )

        return JsonResponse(
            {
                "success": True,
                "message": "Consentimiento informado registrado exitosamente",
                "consent": {
                    "participant_name": participant.name,
                    "participant_email": participant.email,
                    "event_name": event.name,
                    "accepted_at": consent.accepted_at.isoformat(),
                    "consent_version": consent.consent_version,
                    "ip_address": consent.ip_address,
                },
            },
            status=200,
        )

    except Exception as e:
        logger.error(f"Error al registrar consentimiento: {str(e)}")
        return JsonResponse(
            {"error": "Error interno del servidor", "details": str(e)}, status=500
        )


@csrf_exempt
@require_GET
def check_event_consent(request):
    """
    Verifica si existe consentimiento informado para un participante y evento.

    Header requerido:
        Authorization: Bearer {event_key}

    Returns:
        200: Información sobre el estado del consentimiento
        401: Authorization header inválido
        404: Event_key no encontrado
        500: Error del servidor
    """
    try:
        # Validar authorization header
        authorization = request.headers.get("Authorization", "")
        if not authorization.startswith("Bearer "):
            return JsonResponse({"error": "Authorization header inválido"}, status=401)

        parts = authorization.split(" ", 1)
        if len(parts) != 2 or not parts[1]:
            return JsonResponse({"error": "Event_key no proporcionado"}, status=401)

        event_key = parts[1]

        # Obtener participante y evento
        try:
            participant_event = ParticipantEvent.objects.select_related(
                "participant", "event"
            ).get(event_key=event_key)
            participant = participant_event.participant
            event = participant_event.event
        except ParticipantEvent.DoesNotExist:
            return JsonResponse({"error": "Event_key inválido"}, status=404)

        # Buscar consentimiento
        consent = EventConsent.objects.filter(
            participant=participant, event=event
        ).first()

        if consent:
            return JsonResponse(
                {
                    "consentExists": True,
                    "consent": {
                        "accepted_at": consent.accepted_at.isoformat(),
                        "consent_version": consent.consent_version,
                        "participant_name": participant.name,
                        "event_name": event.name,
                    },
                },
                status=200,
            )
        else:
            return JsonResponse(
                {
                    "consentExists": False,
                    "message": "No existe consentimiento registrado. Debe aceptar el consentimiento informado antes de continuar.",
                    "participant": {
                        "name": participant.name,
                        "email": participant.email,
                    },
                    "event": {"name": event.name, "description": event.description},
                },
                status=200,
            )

    except Exception as e:
        logger.error(f"Error al verificar consentimiento: {str(e)}")
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@csrf_exempt
@jwt_required(roles=["admin", "superadmin"])
@require_POST
def unblock_participants(request, event_id):
    """Desbloquea participantes seleccionados de un evento"""
    try:
        event = Event.objects.get(id=event_id)
        data = json.loads(request.body)
        participant_ids = data.get("participant_ids", [])

        if not participant_ids:
            return JsonResponse(
                {"error": "No se especificaron participantes"}, status=400
            )

        # Desbloquear los participantes seleccionados
        unblocked_count = ParticipantEvent.objects.filter(
            event=event, participant_id__in=participant_ids
        ).update(is_blocked=False)

        return JsonResponse(
            {
                "success": True,
                "message": f"{unblocked_count} participante(s) desbloqueado(s) exitosamente",
            }
        )

    except Event.DoesNotExist:
        return JsonResponse({"error": "Evento no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
